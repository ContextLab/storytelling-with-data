"""T011: Read the three source `.xlsx` files into tidy frames + metadata.

Captures `original_header` for every column, infers `inferred_type` per
column, and asserts the observed `source_row_count` matches data-model.md
§2 within ±1.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Workbook / Sheet / Column metadata
# ---------------------------------------------------------------------------


WORKBOOK_FILES: dict[str, str] = {
    "rtu_2025_26": "2025-2026 All RTU Data .xlsx",
    "hs_core": "High School Core Measures 2022-2025.xlsx",
    "ms_core": "Middle School Core Measures Data 2022 - 2025.xlsx",
}


WORKBOOK_TITLES: dict[str, str] = {
    "rtu_2025_26": "RTU Follow-Up Surveys 2025–26",
    "hs_core": "High School Core Measures 2022–2025",
    "ms_core": "Middle School Core Measures 2022–2025",
}


WORKBOOK_KIND: dict[str, str] = {
    "rtu_2025_26": "rtu_event_followup",
    "hs_core": "core_measures",
    "ms_core": "core_measures",
}


# Expected source row counts per data-model.md §2 (the row right after the header).
# Tolerance: ±1.
EXPECTED_ROW_COUNTS: dict[str, dict[str, int]] = {
    "rtu_2025_26": {
        "Kickoff Follow Up": 465,
        "Event 2 Follow Up": 491,
        "Event 3 Follow Up": 443,
        "Event 4 Follow Up": 294,
        "Event 5 Follow up": 157,
    },
    "hs_core": {
        "2022": 472,
        "2023": 229,
        "2025": 563,
    },
    "ms_core": {
        "2022": 417,
        "2023": 305,
        "2024": 419,
        "2025": 264,
    },
}


# RTU sheet → (event_id, ordinal, period_label)
RTU_EVENT_META: dict[str, tuple[str, int, str]] = {
    "Kickoff Follow Up": ("kickoff", 1, "Kickoff"),
    "Event 2 Follow Up": ("event_2", 2, "Event 2"),
    "Event 3 Follow Up": ("event_3", 3, "Event 3"),
    "Event 4 Follow Up": ("event_4", 4, "Event 4"),
    "Event 5 Follow up": ("event_5", 5, "Event 5"),
}


@dataclass
class SheetData:
    """One sheet's tidy data + per-column metadata."""

    workbook_id: str
    sheet_id: str
    display_name: str
    period: int | str | None
    period_label: str
    source_row_count: int
    usable_row_count: int
    df: pd.DataFrame  # rows × columns; column labels = original_header
    headers: list[str]  # original header text in column order
    column_ids: list[str]  # sheet-scoped slugs
    inferred_types: list[str]  # one per column
    missing_pct: list[float]
    top_values: list[list[dict[str, Any]]]
    rtu_session: list[str | None]  # 'gerety' | 'community_building' | 'makeup' | None
    rtu_role: list[str | None]     # 'student' | 'caregiver' | None
    rtu_instance: list[int | None]  # 1..6 for repeated questions; None for non-repeated


def slugify(text: str) -> str:
    """Stable slug from arbitrary header text."""
    s = re.sub(r"[^a-zA-Z0-9]+", "_", str(text).strip().lower()).strip("_")
    return s or "x"


# RTU session/role mapping per the user's domain note:
#   instances 1, 2  → "gerety" session  (1 = student, 2 = caregiver)
#   instances 3, 4  → "community building" session
#   instances 5, 6  → "makeup" session
# A repeated question with no number suffix is implicitly instance 1 (student).
# A column whose base text never repeats anywhere in the sheet has no session/role.
_INSTANCE_TO_SESSION_ROLE = {
    1: ("gerety", "student"),
    2: ("gerety", "caregiver"),
    3: ("community_building", "student"),
    4: ("community_building", "caregiver"),
    5: ("makeup", "student"),
    6: ("makeup", "caregiver"),
}

_HEADER_NUMBER_RE = re.compile(r"^(.*?)(?:\s+(\d+))?$")


def _split_header(h: str) -> tuple[str, int | None]:
    """Return (base_header, instance) where instance is the trailing number
    (1 if missing) — caller must determine whether the question repeats."""
    m = _HEADER_NUMBER_RE.match(str(h).strip())
    if not m:
        return str(h).strip(), None
    base = m.group(1).strip()
    inst = int(m.group(2)) if m.group(2) else None
    return base, inst


def derive_rtu_session_role(headers: list[str]) -> tuple[list[str | None], list[str | None], list[int | None]]:
    """For RTU sheets: assign each column a (session, role, instance).

    Rules:
      - A header that ends in " N" (N >= 2) is instance N.
      - The 'base' (no suffix) version of a repeated header is implicitly instance 1.
      - A header that has no repeated counterparts has no session/role/instance.
    """
    bases: list[str] = []
    nums: list[int | None] = []
    for h in headers:
        base, n = _split_header(h)
        bases.append(base.lower())
        nums.append(n)
    # count occurrences of each base
    counts: dict[str, int] = {}
    for b in bases:
        counts[b] = counts.get(b, 0) + 1

    sessions: list[str | None] = []
    roles: list[str | None] = []
    instances: list[int | None] = []
    for base, n in zip(bases, nums):
        if counts.get(base, 0) <= 1:
            # not a repeated question
            sessions.append(None); roles.append(None); instances.append(None)
            continue
        inst = n if n is not None else 1
        sr = _INSTANCE_TO_SESSION_ROLE.get(inst)
        if sr is None:
            sessions.append(None); roles.append(None); instances.append(inst)
        else:
            sessions.append(sr[0]); roles.append(sr[1]); instances.append(inst)
    return sessions, roles, instances


def infer_type(series: pd.Series) -> str:
    """Heuristic column type inference.

    Order: empty → datetime → numeric → freetext (long strings) → categorical/ordinal.
    """
    non_null = series.dropna()
    if len(non_null) == 0:
        return "empty"

    # Datetime?
    if pd.api.types.is_datetime64_any_dtype(non_null):
        return "datetime"
    sample = non_null.iloc[0]
    if hasattr(sample, "year"):
        return "datetime"

    # Numeric?
    numeric_share = pd.to_numeric(non_null, errors="coerce").notna().mean()
    if numeric_share > 0.95:
        # treat low-cardinality numerics as categorical/ordinal anyway
        if non_null.nunique() <= 6:
            return "ordinal"
        return "numeric"

    # Freetext: long avg length, high cardinality
    str_lengths = non_null.astype(str).str.len()
    avg_len = float(str_lengths.mean())
    cardinality = non_null.nunique()
    if avg_len >= 30 and cardinality / max(len(non_null), 1) > 0.5:
        return "freetext"

    # Categorical: low cardinality
    if cardinality <= 12:
        return "categorical"
    # Long-tail short strings → freetext
    if avg_len >= 20:
        return "freetext"
    return "categorical"


def top_values_for(series: pd.Series, k: int = 5) -> list[dict[str, Any]]:
    counts = series.dropna().astype(str).value_counts().head(k)
    return [{"value": v, "count": int(c)} for v, c in counts.items()]


def read_workbooks(data_dir: Path) -> list[SheetData]:
    """Read all three workbooks and return a flat list of SheetData."""
    out: list[SheetData] = []
    for wb_id, fname in WORKBOOK_FILES.items():
        path = data_dir / fname
        if not path.exists():
            raise FileNotFoundError(f"missing workbook: {path}")
        wb = load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = _read_sheet(wb_id, sheet_name, ws)
            _verify_row_count(sheet_data)
            out.append(sheet_data)
    return out


def _read_sheet(wb_id: str, sheet_name: str, ws) -> SheetData:
    # Capture original headers verbatim
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        headers.append(str(v) if v is not None else f"<empty col {col}>")

    # Read body via pandas (faster + handles types). We use openpyxl-loaded
    # values for headers because pandas would dedupe them.
    df = pd.read_excel(
        ws.parent.path if hasattr(ws.parent, "path") else None,
        sheet_name=sheet_name,
        engine="openpyxl",
        header=0,
    ) if False else _df_from_ws(ws, headers)

    # source_row_count: total rows including header, matching the values
    # from data-model.md §2 (which were derived from openpyxl's max_row).
    source_row_count = ws.max_row

    # usable = rows where at least one non-school/grade cell is non-null
    usable_row_count = int(df.dropna(how="all").shape[0])

    # Per-column stats — positional access (df may have duplicate header labels)
    inferred_types: list[str] = []
    missing_pcts: list[float] = []
    top_vals: list[list[dict[str, Any]]] = []
    for col_idx in range(len(headers)):
        if col_idx < df.shape[1]:
            ser = df.iloc[:, col_idx]
        else:
            ser = pd.Series(dtype=object)
        inferred_types.append(infer_type(ser))
        if len(ser) == 0:
            missing_pcts.append(100.0)
        else:
            missing_pcts.append(float(ser.isna().mean() * 100.0))
        top_vals.append(top_values_for(ser))

    column_ids = _make_column_ids(wb_id, sheet_name, headers)

    period, period_label = _period_for(wb_id, sheet_name)
    sheet_id = _sheet_id_for(wb_id, sheet_name)

    # RTU sheets: derive session/role/instance per column
    if wb_id == "rtu_2025_26":
        rtu_sessions, rtu_roles, rtu_instances = derive_rtu_session_role(headers)
    else:
        rtu_sessions = [None] * len(headers)
        rtu_roles = [None] * len(headers)
        rtu_instances = [None] * len(headers)

    return SheetData(
        workbook_id=wb_id,
        sheet_id=sheet_id,
        display_name=sheet_name,
        period=period,
        period_label=period_label,
        source_row_count=source_row_count,
        usable_row_count=usable_row_count,
        df=df,
        headers=headers,
        column_ids=column_ids,
        inferred_types=inferred_types,
        missing_pct=missing_pcts,
        top_values=top_vals,
        rtu_session=rtu_sessions,
        rtu_role=rtu_roles,
        rtu_instance=rtu_instances,
    )


def _df_from_ws(ws, headers: list[str]) -> pd.DataFrame:
    """Build a DataFrame from an openpyxl worksheet using captured headers
    as column labels (preserves duplicates with disambiguation)."""
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row) + [None] * max(0, len(headers) - len(row)))
    # Disambiguate duplicate headers by appending position suffix
    seen: dict[str, int] = {}
    cols: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            cols.append(f"{h}#{seen[h]}")
        else:
            seen[h] = 0
            cols.append(h)
    df = pd.DataFrame(rows, columns=cols)
    # Restore the original (possibly duplicated) headers in df.columns —
    # downstream code uses positional access via column_ids, but we keep
    # readable labels for debugging.
    df.columns = headers
    return df


def _make_column_ids(wb_id: str, sheet_name: str, headers: Iterable[str]) -> list[str]:
    sid = _sheet_id_for(wb_id, sheet_name)
    return [f"{sid}.q{idx:02d}_{slugify(h)[:40]}" for idx, h in enumerate(headers)]


def _sheet_id_for(wb_id: str, sheet_name: str) -> str:
    if wb_id == "rtu_2025_26":
        meta = RTU_EVENT_META.get(sheet_name)
        if meta is not None:
            return f"{wb_id}.{meta[0]}"
        return f"{wb_id}.{slugify(sheet_name)}"
    # core_measures: sheet name is the year
    return f"{wb_id}.{slugify(sheet_name)}"


def _period_for(wb_id: str, sheet_name: str) -> tuple[int | str | None, str]:
    if wb_id == "rtu_2025_26":
        meta = RTU_EVENT_META.get(sheet_name)
        if meta is not None:
            return meta[1], meta[2]
        return None, sheet_name
    # core_measures: parse year
    try:
        year = int(sheet_name)
        return year, str(year)
    except ValueError:
        return None, sheet_name


def _verify_row_count(sheet: SheetData) -> None:
    expected = EXPECTED_ROW_COUNTS.get(sheet.workbook_id, {}).get(sheet.display_name)
    if expected is None:
        return
    if abs(sheet.source_row_count - expected) > 1:
        raise AssertionError(
            f"row-count mismatch for {sheet.workbook_id}/{sheet.display_name}: "
            f"observed {sheet.source_row_count}, expected ~{expected} (±1)"
        )


def event_meta_for_sheet(sheet: SheetData) -> tuple[str, int, str] | None:
    """Return (event_id, ordinal, label) for an RTU sheet, else None."""
    if sheet.workbook_id != "rtu_2025_26":
        return None
    return RTU_EVENT_META.get(sheet.display_name)
